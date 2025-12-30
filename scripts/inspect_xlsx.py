import sys
try:
    import openpyxl
except Exception as e:
    print('MISSING_OPENPYXL')
    sys.exit(2)
wb = openpyxl.load_workbook('OSSRequestForm-v4.xlsx')
print('SHEETS:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('\n--- SHEET:', name)
    for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=8, values_only=True):
        print(row)
